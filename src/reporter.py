from datetime import datetime
from pathlib import Path

import openpyxl as opxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from src.processing import Processing


class Reporter:
    def __init__(self, df_data: pd.DataFrame, output_folder: str):
        self.df_data = df_data
        self.output_folder = output_folder
        
    def _generate_summary(self) -> pd.DataFrame:
        
        summary_data = {
            "Descrição": [
                "Total de Cobranças (Interno)",
                "Total de Cobranças (CSV)",
                "Valor Líquido Total (CSV)",
                "Total de Glosas",
                "Laudos Renomeados com Sucesso"
            ],
            "Valor": [ 
                len(self.df_data[self.df_data['_merge'] != 'right_only']),
                len(self.df_data[self.df_data['_merge'] != 'left_only']),
                self.df_data['vl_liquido'].apply(Processing.normalize_value_csv_to_float).sum(),
                self.df_data['vl_glosa'].apply(Processing.normalize_value_csv_to_float).sum(),
                self.df_data['arquivo_renomeado'].count()
            ]
        }
        return pd.DataFrame(summary_data)

    def _generate_detailed_data(self) -> pd.DataFrame:
        detailed_data = {
            "ID Cobrança": self.df_data['num_guia'],   
            "Procedimento": self.df_data['procedimento'],   
            "CPF Beneficiário": self.df_data['cpf_beneficiario'],
            "Código TUSS": self.df_data['cod_tuss'],
            "Nome Operadora": self.df_data['nome_operadora'],
            "PDF Renomeado": self.df_data['arquivo_renomeado'],
            "Divergencias": self.df_data['divergencias'],
        }
        return pd.DataFrame(detailed_data)
    
    def _generate_alerts(self) -> pd.DataFrame:
        
        
      df_alerts = self.df_data[self.df_data['divergencias'].notna() & (self.df_data['divergencias'] != "Ok")]
      
      return df_alerts

    def generate_report(self) -> None:
        summary_df = self._generate_summary()
        detailed_df = self._generate_detailed_data()
        alerts_df = self._generate_alerts()
        
        
        wb = opxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "Resumo"
        headers = summary_df.columns.tolist()
        ws_summary.append(headers)
        
        wb_detailed = wb.create_sheet("Detalhes")
        headers_detailed = detailed_df.columns.tolist()
        wb_detailed.append(headers_detailed)
        
        wb_alerts = wb.create_sheet("Alertas")
        headers_alerts = alerts_df.columns.tolist()
        wb_alerts.append(headers_alerts) 
        
        header_font = Font(bold=True, color="DC143C", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")
    

        for row in summary_df.itertuples(index=False): 
            ws_summary.append(row)
            
        for row in detailed_df.itertuples(index=False):
            wb_detailed.append(row)
            
        for row in alerts_df.itertuples(index=False):
            wb_alerts.append(row) 
            
        for row in [ws_summary['1'], wb_detailed['1'], wb_alerts['1']]:
            for cell in row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
            
        for sheet in [ws_summary, wb_detailed, wb_alerts]: 
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
        
                for cell in column:
                    cell.alignment = alignment
            
                    try:
                        val_str = str(cell.value) if cell.value is not None else ""
                        if len(val_str) > max_length:
                            max_length = len(val_str)
                    except:
                        pass
        
                    adjusted_width = (max_length + 4)
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            
             
            
        
        
        path_name = f"relatorio_faturamento_{datetime.now().strftime('%Y-%m')}.xlsx"
        
        output_path = Path(self.output_folder) / path_name
        output_path.parent.mkdir(parents=True, exist_ok=True)

        if output_path.exists():
            output_path.unlink() 

        wb.save(output_path)
        
         

        